from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# --- Configuration ---
VIDEO_DIR = Path("best_media")
THUMB_DIR = Path("thumbnails")
OUTPUT_FILE = "270_video_analysis .xlsx"
THUMB_HEIGHT = 180  # pixels, all thumbnails will have this height

VIDEO_EXTS = {".mp4", ".mov", ".avi", ".mkv", ".webm"}

# --- Step 1: Gather video names and thumbnail paths ---
rows = []
for video_file in VIDEO_DIR.iterdir():
    if video_file.suffix.lower() not in VIDEO_EXTS:
        continue
    thumb_file = THUMB_DIR / f"{video_file.stem}.jpg"
    if not thumb_file.exists():
        continue
    rows.append({
        "Video Name": video_file.name,
        "Thumbnail Path": thumb_file
    })

# --- Step 2: Create Excel workbook ---
wb = Workbook()
ws = wb.active
ws.title = "Videos"

# Header row
ws["A1"] = "Video Name"
ws["B1"] = "Thumbnail"

# Optional: set column A width for video names
ws.column_dimensions['A'].width = 30

# --- Step 3: Add video names and embed thumbnails ---
for idx, row in enumerate(rows, start=2):  # start=2 for header
    # Video name
    ws[f"A{idx}"] = row["Video Name"]

    # Embed thumbnail
    img_path = row["Thumbnail Path"]
    pil_img = PILImage.open(img_path)

    # Calculate width to preserve aspect ratio
    aspect_ratio = pil_img.width / pil_img.height
    target_height = THUMB_HEIGHT
    target_width = int(target_height * aspect_ratio)

    img = XLImage(str(img_path))
    img.width = target_width
    img.height = target_height

    ws.add_image(img, f"B{idx}")

    # Adjust row height (Excel row height roughly = px * 0.75)
    ws.row_dimensions[idx].height = target_height * 0.75

# Optional: adjust thumbnail column width based on largest width
max_width = max(int(THUMB_HEIGHT * (PILImage.open(r["Thumbnail Path"]).width / PILImage.open(r["Thumbnail Path"]).height)) for r in rows)
ws.column_dimensions['B'].width = max_width / 7  # Excel units ≈ px / 7

# --- Step 4: Save workbook ---
wb.save(OUTPUT_FILE)
print(f"Saved {OUTPUT_FILE} with {len(rows)} thumbnails embedded.")
